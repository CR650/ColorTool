"""
Python后端文件处理服务
Flask + openpyxl 实现RSC_Theme文件处理
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import json
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)  # 允许跨域请求

class ThemeProcessor:
    def __init__(self):
        self.mapping_data = self.load_mapping_data()
    
    def load_mapping_data(self):
        """加载颜色映射数据"""
        return [
            {"RC现在的主题通道": "P1", "作用": "地板颜色", "颜色代码": "P1"},
            {"RC现在的主题通道": "P5", "作用": "跳板颜色", "颜色代码": "P2"},
            {"RC现在的主题通道": "G1", "作用": "装饰颜色1", "颜色代码": "G1"},
            {"RC现在的主题通道": "G2", "作用": "装饰颜色2", "颜色代码": "G2"},
            {"RC现在的主题通道": "G3", "作用": "装饰颜色3", "颜色代码": "G3"},
            {"RC现在的主题通道": "G4", "作用": "装饰颜色4", "颜色代码": "G4"},
            {"RC现在的主题通道": "P2", "作用": "地板描边颜色", "颜色代码": "P1-1"},
            {"RC现在的主题通道": "P9", "作用": "地板侧面颜色", "颜色代码": "P1-2"},
            {"RC现在的主题通道": "P6", "作用": "跳板描边颜色", "颜色代码": "P2-1"},
            {"RC现在的主题通道": "P10", "作用": "跳板侧面颜色", "颜色代码": "P2-2"},
            {"RC现在的主题通道": "G5", "作用": "装饰颜色5", "颜色代码": "G5"},
            {"RC现在的主题通道": "G6", "作用": "装饰颜色6", "颜色代码": "G6"},
            {"RC现在的主题通道": "G7", "作用": "装饰颜色7", "颜色代码": "G7"}
        ]
    
    def find_color_value(self, source_data, color_code):
        """在源数据中查找颜色值"""
        for row in source_data:
            # 尝试多种字段名称匹配
            color_code_fields = ['颜色代码', 'colorCode', 'code', '代码']
            color_value_fields = ['16进制值', '颜色值', 'hex', 'HEX', 'hexValue']
            
            row_color_code = None
            for field in color_code_fields:
                if field in row and row[field]:
                    row_color_code = str(row[field]).strip()
                    break
            
            if row_color_code and row_color_code.upper() == color_code.upper():
                # 查找颜色值
                for field in color_value_fields:
                    if field in row and row[field]:
                        color_value = str(row[field]).strip().upper()
                        # 移除#号并验证格式
                        if color_value.startswith('#'):
                            color_value = color_value[1:]
                        if len(color_value) == 6 and all(c in '0123456789ABCDEF' for c in color_value):
                            return color_value
                
                # 尝试RGB值
                if all(f in row for f in ['R值', 'G值', 'B值']):
                    try:
                        r = int(row['R值'])
                        g = int(row['G值'])
                        b = int(row['B值'])
                        return f"{r:02X}{g:02X}{b:02X}"
                    except:
                        pass
        
        return None
    
    def process_theme(self, rsc_data, source_data, theme_name):
        """处理主题数据"""
        # 转换为DataFrame便于处理
        df = pd.DataFrame(rsc_data[1:], columns=rsc_data[0])
        
        # 查找或创建主题行
        theme_row_index = None
        if 'notes' in df.columns:
            existing_theme = df[df['notes'] == theme_name]
            if not existing_theme.empty:
                theme_row_index = existing_theme.index[0]
            else:
                # 创建新行
                new_row = {col: '' for col in df.columns}
                new_row['notes'] = theme_name
                if 'id' in df.columns:
                    max_id = df['id'].apply(lambda x: int(x) if str(x).isdigit() else 0).max()
                    new_row['id'] = str(max_id + 1)
                
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                theme_row_index = len(df) - 1
        
        # 更新颜色数据
        updated_colors = []
        summary = {"total": 0, "updated": 0, "notFound": 0, "errors": []}
        
        for mapping in self.mapping_data:
            rc_channel = mapping['RC现在的主题通道']
            color_code = mapping['颜色代码']
            
            if rc_channel not in df.columns:
                summary["errors"].append(f"未找到列: {rc_channel}")
                continue
            
            summary["total"] += 1
            
            # 查找颜色值
            color_value = self.find_color_value(source_data, color_code)
            
            if color_value:
                df.at[theme_row_index, rc_channel] = color_value
                updated_colors.append({
                    "channel": rc_channel,
                    "colorCode": color_code,
                    "value": color_value,
                    "isDefault": False
                })
                summary["updated"] += 1
            else:
                # 使用默认值
                df.at[theme_row_index, rc_channel] = "FFFFFF"
                updated_colors.append({
                    "channel": rc_channel,
                    "colorCode": color_code,
                    "value": "FFFFFF",
                    "isDefault": True
                })
                summary["notFound"] += 1
        
        # 转换回数组格式
        result_data = [df.columns.tolist()] + df.values.tolist()
        
        return {
            "success": True,
            "data": result_data,
            "updatedColors": updated_colors,
            "summary": summary,
            "themeName": theme_name,
            "rowIndex": theme_row_index
        }

# 创建处理器实例
processor = ThemeProcessor()

@app.route('/api/process-theme', methods=['POST'])
def process_theme():
    """处理主题数据的API端点"""
    try:
        # 获取上传的文件和数据
        rsc_file = request.files.get('rsc_file')
        source_file = request.files.get('source_file')
        theme_name = request.form.get('theme_name')
        
        if not all([rsc_file, source_file, theme_name]):
            return jsonify({"error": "缺少必要的文件或参数"}), 400
        
        # 读取RSC文件
        rsc_workbook = openpyxl.load_workbook(rsc_file)
        rsc_sheet = rsc_workbook.active
        rsc_data = [[cell.value for cell in row] for row in rsc_sheet.iter_rows()]
        
        # 读取源数据文件
        if source_file.filename.endswith('.xlsx'):
            source_df = pd.read_excel(source_file, sheet_name='完整配色表' if '完整配色表' in pd.ExcelFile(source_file).sheet_names else 0)
        else:
            source_df = pd.read_excel(source_file)
        
        source_data = source_df.to_dict('records')
        
        # 处理主题
        result = processor.process_theme(rsc_data, source_data, theme_name)
        
        if result["success"]:
            # 创建新的工作簿
            new_workbook = openpyxl.Workbook()
            new_sheet = new_workbook.active
            
            # 写入数据
            for row in result["data"]:
                new_sheet.append(row)
            
            # 保存到内存
            output = io.BytesIO()
            new_workbook.save(output)
            output.seek(0)
            
            # 返回文件
            return send_file(
                output,
                as_attachment=True,
                download_name=f'RSC_Theme_Updated_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify(result), 400
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查端点"""
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat()})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
